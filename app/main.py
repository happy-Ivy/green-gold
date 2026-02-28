from datetime import datetime, timedelta
from pathlib import Path
import os
from io import BytesIO

from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlmodel import Session, select
from openpyxl import Workbook
from dotenv import load_dotenv

from .db import init_db, get_session
from .models import User, EmailOTP, TransactionCode, PointLog
from .security import gen_otp, hash_otp, gen_code
from .emailer import send_otp_email

load_dotenv()

# ✅ 只保留一個 app
app = FastAPI(title="Green Points")
app.add_middleware(SessionMiddleware, secret_key="CHANGE_ME_TO_A_LONG_RANDOM_SECRET")

# ✅ 用絕對路徑掛載 static（不怕你從哪個目錄啟動）
BASE_DIR = Path(__file__).resolve().parent   # app/
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")

# ✅ templates 也用一致路徑
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

OTP_TTL_MINUTES = 10
OTP_MAX_ATTEMPTS = 5
def admin_whitelist() -> set[str]:
    raw = os.getenv("ADMIN_EMAILS", "")
    return {e.strip().lower() for e in raw.split(",") if e.strip()}


def require_login(request: Request) -> dict:
    user_id = request.session.get("user_id")
    role = request.session.get("role")
    if not user_id or not role:
        raise HTTPException(status_code=401, detail="Not logged in")
    return {"user_id": user_id, "role": role}

@app.on_event("startup")
def on_startup():
    init_db()

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/otp", response_class=HTMLResponse)
def otp_page(request: Request):
    pending = request.session.get("pending_email")
    if not pending:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("otp.html", {"request": request, "email": pending})
@app.post("/request-otp")
async def request_otp(
    request: Request,
    email: str = Form(...),
    role: str = Form(...),  # "user" or "merchant"
    session: Session = Depends(get_session),
):
    email = email.strip().lower()

    # 允許 user / merchant / admin
    if role not in ("user", "merchant", "admin"):
        raise HTTPException(status_code=400, detail="Invalid role")

    # admin 必須白名單
    if role == "admin" and email not in admin_whitelist():
        raise HTTPException(status_code=403, detail="Not allowed to login as admin")

    # 查使用者
    user = session.exec(select(User).where(User.email == email)).first()

    if not user:
        # 新用戶：不允許自動建立 admin
        if role == "admin":
            raise HTTPException(status_code=403, detail="Admin account must be pre-approved")
        user = User(email=email, role=role)
        session.add(user)
        session.commit()
        session.refresh(user)
    else:
        # 已存在：角色固定，不允許從登入頁切換
        role = user.role

    # 已存在但角色是 admin：仍要白名單保護（避免資料庫有 admin 但環境變數未設定）
    if role == "admin" and email not in admin_whitelist():
        raise HTTPException(status_code=403, detail="Admin not allowed")

    # 產生 OTP 並寫入 DB
    otp = gen_otp()
    salt = gen_code(10)
    otp_hash = hash_otp(otp, salt)
    expires_at = datetime.utcnow() + timedelta(minutes=OTP_TTL_MINUTES)

    rec = EmailOTP(email=email, otp_hash=f"{salt}:{otp_hash}", expires_at=expires_at)
    session.add(rec)
    session.commit()

    # 寄信（因為在 async function 裡，所以 await OK）
    await send_otp_email(email, otp)

    # 設定 session pending email
    request.session["pending_email"] = email

    return RedirectResponse(url="/otp", status_code=303)

@app.post("/verify-otp")
def verify_otp(
    request: Request,
    otp: str = Form(...),
    session: Session = Depends(get_session),
):
    email = request.session.get("pending_email")
    if not email:
        return RedirectResponse(url="/login", status_code=303)

    # Latest unused OTP
    rec = session.exec(
        select(EmailOTP)
        .where(EmailOTP.email == email)
        .where(EmailOTP.is_used == False)
        .order_by(EmailOTP.id.desc())
    ).first()

    if not rec:
        raise HTTPException(400, "No OTP found. Request again.")

    if datetime.utcnow() > rec.expires_at:
        raise HTTPException(400, "OTP expired. Request again.")

    if rec.attempts >= OTP_MAX_ATTEMPTS:
        raise HTTPException(429, "Too many attempts. Request a new OTP.")

    salt, hashed = rec.otp_hash.split(":", 1)
    if hash_otp(otp.strip(), salt) != hashed:
        rec.attempts += 1
        session.add(rec)
        session.commit()
        raise HTTPException(400, "Invalid OTP")

    rec.is_used = True
    session.add(rec)

    user = session.exec(select(User).where(User.email == email)).first()
    if not user:
        raise HTTPException(400, "User not found")

    session.commit()

    request.session.pop("pending_email", None)
    request.session["user_id"] = user.id
    request.session["role"] = user.role

    return RedirectResponse(url="/home", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)

@app.get("/rules", response_class=HTMLResponse)
def rules_page(request: Request):
    require_login(request)  # ✅ 登入後才能看，若你想不登入也能看就拿掉這行
    return templates.TemplateResponse("rules.html", {"request": request})

@app.get("/home", response_class=HTMLResponse)
def home(
    request: Request,
    session: Session = Depends(get_session),
):
    auth = require_login(request)
    user = session.get(User, auth["user_id"])
    if not user:
        request.session.clear()
        return RedirectResponse(url="/login", status_code=303)

    # ✅ Admin dashboard
    if user.role == "admin":
        users = session.exec(select(User).order_by(User.id.desc())).all()
        codes = session.exec(select(TransactionCode).order_by(TransactionCode.id.desc())).all()
        logs = session.exec(select(PointLog).order_by(PointLog.id.desc())).all()
        return templates.TemplateResponse(
            "admin_home.html",
            {"request": request, "user": user, "users": users, "codes": codes, "logs": logs},
        )

    # ✅ Merchant dashboard
    if user.role == "merchant":
        codes = session.exec(
            select(TransactionCode)
            .where(TransactionCode.merchant_id == user.id)
            .order_by(TransactionCode.id.desc())
            .limit(50)
        ).all()
        return templates.TemplateResponse(
            "merchant_home.html",
            {"request": request, "user": user, "codes": codes},
        )

    # ✅ User dashboard (default)
    logs = session.exec(
        select(PointLog)
        .where(PointLog.user_id == user.id)
        .order_by(PointLog.id.desc())
        .limit(50)
    ).all()
    return templates.TemplateResponse(
        "user_home.html",
        {"request": request, "user": user, "logs": logs},
    )
@app.get("/admin/export")
def admin_export(request: Request, session: Session = Depends(get_session)):
    auth = require_login(request)
    user = session.get(User, auth["user_id"])
    if not user or user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    wb = Workbook()

    # Sheet 1: users
    ws = wb.active
    ws.title = "users"
    ws.append(["id", "email", "role", "green_points", "created_at"])
    users = session.exec(select(User).order_by(User.id.asc())).all()
    for u in users:
        ws.append([u.id, u.email, u.role, u.green_points, str(u.created_at)])

    # Sheet 2: codes
    ws2 = wb.create_sheet("codes")
    ws2.append(["id", "code", "merchant_id", "points", "is_used", "created_at", "used_at", "used_by_user_id"])
    codes = session.exec(select(TransactionCode).order_by(TransactionCode.id.asc())).all()
    for c in codes:
        ws2.append([c.id, c.code, c.merchant_id, c.points, c.is_used, str(c.created_at), str(c.used_at), c.used_by_user_id])

    # Sheet 3: logs
    ws3 = wb.create_sheet("logs")
    ws3.append(["id", "user_id", "merchant_id", "points", "code", "created_at"])
    logs = session.exec(select(PointLog).order_by(PointLog.id.asc())).all()
    for l in logs:
        ws3.append([l.id, l.user_id, l.merchant_id, l.points, l.code, str(l.created_at)])

    # Output
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"green_points_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/merchant/create", response_class=HTMLResponse)
def merchant_create_page(request: Request):
    auth = require_login(request)
    if auth["role"] != "merchant":
        raise HTTPException(403, "Merchant only")
    return templates.TemplateResponse("create_code.html", {"request": request})

@app.post("/merchant/create")
def merchant_create_code(
    request: Request,
    points: int = Form(...),
    session: Session = Depends(get_session),
):
    auth = require_login(request)
    if auth["role"] != "merchant":
        raise HTTPException(403, "Merchant only")
    if points <= 0 or points > 100000:
        raise HTTPException(400, "Invalid points")

    # Generate unique code
    for _ in range(10):
        code = gen_code(8)
        exists = session.exec(select(TransactionCode).where(TransactionCode.code == code)).first()
        if not exists:
            break
    else:
        raise HTTPException(500, "Failed to generate code")

    rec = TransactionCode(code=code, merchant_id=auth["user_id"], points=points)
    session.add(rec)
    session.commit()

    return RedirectResponse(url="/home", status_code=303)

@app.get("/redeem", response_class=HTMLResponse)
def redeem_page(request: Request):
    auth = require_login(request)
    if auth["role"] != "user":
        raise HTTPException(403, "User only")
    return templates.TemplateResponse("redeem.html", {"request": request})

@app.post("/redeem")
def redeem(
    request: Request,
    code: str = Form(...),
    session: Session = Depends(get_session),
):
    auth = require_login(request)
    if auth["role"] != "user":
        raise HTTPException(403, "User only")

    code = code.strip().upper()
    rec = session.exec(select(TransactionCode).where(TransactionCode.code == code)).first()
    if not rec:
        raise HTTPException(400, "Code not found")
    if rec.is_used:
        raise HTTPException(400, "Code already used")

    user = session.get(User, auth["user_id"])
    if not user:
        raise HTTPException(400, "User not found")

    # Apply points
    user.green_points += rec.points
    rec.is_used = True
    rec.used_at = datetime.utcnow()
    rec.used_by_user_id = user.id

    log = PointLog(user_id=user.id, merchant_id=rec.merchant_id, points=rec.points, code=rec.code)
    session.add(user)
    session.add(rec)
    session.add(log)
    session.commit()

    return RedirectResponse(url="/home", status_code=303)
